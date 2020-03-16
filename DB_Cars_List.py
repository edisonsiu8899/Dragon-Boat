#from Download_Attendance_Sheet import name #COMMENT OUT FOR GSHEETS
import string
import random
from openpyxl import load_workbook
#import facebook
import datetime
from datetime import date
import time
import smtplib
from facepy import GraphAPI
from string import ascii_uppercase

def make_cars(Saturday, Sunday):
    #global randmess
    #randmess = "Hello World"
    global drivers

    largebooty = []
    smallbooty = []
    dissapointments = []

    #wb = load_workbook(name + '.xlsx') #COMMENT OUT FOR GSHEETS
    wb = load_workbook('C:\\Users\\ediso\\Desktop\\Dragonboat Spring 2020 Attendance.xlsx')
    ws = wb.active

    alphabet = list(string.ascii_uppercase)
    for cell in alphabet[1:]:
        date = str(ws[cell+'1'].value)[0:10] # get just date portion of string in this format '2019-01-25'
        # print(date)
        if Saturday == date:
            satPractice = ws[cell + '1']
            satIndex = cell
        if Sunday == date:
            sunPractice = ws[cell + '1']
            sunIndex = cell

    attendanceList = list()
    # index 0 is satGoingList
    # index 1 is satEarlyList
    # index 2 is satKGList
    # index 3 is satKGEarlyList

    # index 4 is sunGoingList
    # index 5 is sunEarlyList
    # index 6 is sunKGList
    # index 7 is sunKGEarlyList
    for i in range(8):
        attendanceList.append(list())

    for row in range(2, 29):
        satMemberResponse = satIndex + '{}'.format(row) # response of member for Saturday practice
        sunMemberReponse = sunIndex + '{}'.format(row) # response of member for Sunday practice
        member = ws['A' + str(row)].value
        if ws[satMemberResponse].value == 1:
            attendanceList[0].append(member)
        elif ws[satMemberResponse].value == 'E':
            attendanceList[1].append(member)
        elif ws[satMemberResponse].value == 'K':
            attendanceList[2].append(member)
        elif ws[satMemberResponse].value == 'KE':
            attendanceList[3].append(member)
        if ws[sunMemberReponse].value == 1:
            attendanceList[4].append(member)
        elif ws[sunMemberReponse].value == 'E':
            attendanceList[5].append(member)
        elif ws[sunMemberReponse].value == 'K':
            attendanceList[6].append(member)
        elif ws[sunMemberReponse].value == 'KE':
            attendanceList[7].append(member)

    for i in range(8):
        random.shuffle(attendanceList[i])
    print("\nSaturday Going List: ", attendanceList[0])
    print("Saturday Early Car List:", attendanceList[1])
    print("Saturday KG List:", attendanceList[2])
    print("Saturday KG and Early Car List:", attendanceList[3])
    print("Number of People Going Saturday: " + str(len(attendanceList[0]) + len(attendanceList[1]) +
                                                    len(attendanceList[2]) + len(attendanceList[3])))
    print("Number of people that don't need early car Saturday: " + str(len(attendanceList[0]) + len(attendanceList[2])))
    print("Number of people going early Saturday: " + str(len(attendanceList[1]) + len(attendanceList[3])))
    print("\nSunday Going List: ", attendanceList[4])
    print("Sunday Early Car List:", attendanceList[5])
    print("Sunday KG List:", attendanceList[6])
    print("Sunday KG and Early Car List:", attendanceList[7])
    print("Number of People Going Sunday: " + str(len(attendanceList[4]) + len(attendanceList[5]) +
                                                  len(attendanceList[6]) + len(attendanceList[7])))
    print("Number of people that don't need early car Sunday: " + str(len(attendanceList[4]) + len(attendanceList[6])))
    print("Number of people going early Sunday: " + str(len(attendanceList[5]) + len(attendanceList[7])))

    drivers = ['Amanda', 'Jackie', 'Serena', 'Rain', 'Camila']
    shuffled_drivers = random.sample(drivers, len(drivers))
    i = 0
    print("\n")
    cars_going = {}
    leftover = [] #Leftover people is a combination of people who don't have a spot,
    # and will likely have to discuss among the car what they want to do if they want to leave early or not
    maven_sat = []

    sat_reg_going_drivers = []
    sat_KG_going_drivers = []
    sat_reg_leaving_drivers = []
    sat_early_leaving_drivers = []

    #Saturday Regular
    for i in range(len(shuffled_drivers)):
        if(shuffled_drivers[i] in attendanceList[0]):
            cars_going[shuffled_drivers[i]] = []
            attendanceList[0].remove(shuffled_drivers[i])
            if(len(attendanceList[0]) < 4):
                for j in range(len(attendanceList[0])):
                    choice = random.choice(attendanceList[0])
                    if(choice not in shuffled_drivers):
                        cars_going[shuffled_drivers[i]].append(choice)
                        attendanceList[0].remove(choice)
            else:
                for j in range(4):
                    while(True):
                        choice = random.choice(attendanceList[0])
                        if(choice not in shuffled_drivers):
                            break
                    cars_going[shuffled_drivers[i]].append(choice)
                    attendanceList[0].remove(choice)
            if(len(cars_going[shuffled_drivers[i]]) > 1):
                sat_reg_going_drivers.append(shuffled_drivers[i])
                sat_reg_leaving_drivers.append(shuffled_drivers[i])
                print("Saturday Driver:", shuffled_drivers[i], "|", cars_going[shuffled_drivers[i]])
    if(len(attendanceList[0]) > 0):
        for k in range(len(attendanceList[0])):
            leftover.append(attendanceList[0][k])

    #Saturday Early
    for i in range(len(shuffled_drivers)):
        if(shuffled_drivers[i] in attendanceList[1]):
            cars_going[shuffled_drivers[i]] = []
            attendanceList[1].remove(shuffled_drivers[i])
            if(len(attendanceList[1]) < 4):
                for j in range(len(attendanceList[1])):
                    choice = random.choice(attendanceList[1])
                    if(choice not in shuffled_drivers):
                        cars_going[shuffled_drivers[i]].append(choice)
                        attendanceList[1].remove(choice)
            else:
                for j in range(4):
                    while(True):
                        choice = random.choice(attendanceList[1])
                        if(choice not in shuffled_drivers):
                            break
                    cars_going[shuffled_drivers[i]].append(choice)
                    attendanceList[1].remove(choice)
            if(len(cars_going[shuffled_drivers[i]]) > 1):
                sat_reg_going_drivers.append(shuffled_drivers[i])
                sat_early_leaving_drivers.append(shuffled_drivers[i])
                print("Saturday Early Driver:", shuffled_drivers[i], "|", cars_going[shuffled_drivers[i]])
    if(len(attendanceList[1]) > 0):
        for k in range(len(attendanceList[1])):
            leftover.append(attendanceList[1][k])

    # Saturday KG
    for i in range(len(shuffled_drivers)):
        if (shuffled_drivers[i] in attendanceList[2]):
            cars_going[shuffled_drivers[i]] = []
            attendanceList[2].remove(shuffled_drivers[i])
            if (len(attendanceList[2]) < 4):
                for j in range(len(attendanceList[2])):
                    choice = random.choice(attendanceList[2])
                    if (choice not in shuffled_drivers):
                        cars_going[shuffled_drivers[i]].append(choice)
                        attendanceList[2].remove(choice)
            else:
                for j in range(4):
                    while (True):
                        choice = random.choice(attendanceList[2])
                        if (choice not in shuffled_drivers):
                            break
                    cars_going[shuffled_drivers[i]].append(choice)
                    attendanceList[2].remove(choice)
            if(len(cars_going[shuffled_drivers[i]]) > 1):
                sat_KG_going_drivers.append(shuffled_drivers[i])
                sat_reg_leaving_drivers.append(shuffled_drivers[i])
                print("Saturday KG Driver:", shuffled_drivers[i], "|", cars_going[shuffled_drivers[i]])
    if(len(attendanceList[2]) > 0):
        for k in range(len(attendanceList[2])):
            leftover.append(attendanceList[2][k])

    # Saturday KG/Early
    for i in range(len(shuffled_drivers)):
        if (shuffled_drivers[i] in attendanceList[3]):
            cars_going[shuffled_drivers[i]] = []
            attendanceList[3].remove(shuffled_drivers[i])
            if (len(attendanceList[3]) < 4):
                for j in range(len(attendanceList[3])):
                    choice = random.choice(attendanceList[3])
                    if (choice not in shuffled_drivers):
                        cars_going[shuffled_drivers[i]].append(choice)
                        attendanceList[3].remove(choice)
            else:
                for j in range(4):
                    while(True):
                        choice = random.choice(attendanceList[3])
                        if (choice not in shuffled_drivers):
                            break
                    cars_going[shuffled_drivers[i]].append(choice)
                    attendanceList[3].remove(choice)
            if(len(cars_going[shuffled_drivers[i]]) > 1):
                sat_KG_going_drivers.append(shuffled_drivers[i])
                sat_early_leaving_drivers.append(shuffled_drivers[i])
                print("Saturday KG Early Driver:", shuffled_drivers[i], "|", cars_going[shuffled_drivers[i]])
    if(len(attendanceList[3]) > 0):
        for k in range(len(attendanceList[3])):
            leftover.append(attendanceList[3][k])

    for i in cars_going:
        if(len(cars_going[i]) == 0):
            leftover.append(i)
        if (len(cars_going[i]) == 1):
            leftover.append(i)
            for k in cars_going[i]:
                leftover.append(k)

    for i in range(len(shuffled_drivers)):
        if(shuffled_drivers[i] in leftover):
            cars_going[shuffled_drivers[i]] = []
            leftover.remove(shuffled_drivers[i])
            if(len(leftover) < 4):
                for j in range(len(leftover)):
                    choice = random.choice(leftover)
                    if(choice not in shuffled_drivers):
                        cars_going[shuffled_drivers[i]].append(choice)
                        leftover.remove(choice)
            else:
                for j in range(4):
                    while(True):
                        choice = random.choice(leftover)
                        if(choice not in shuffled_drivers):
                            break
                    cars_going[shuffled_drivers[i]].append(choice)
                    leftover.remove(choice)
            if(len(cars_going[shuffled_drivers[i]]) > 1):
                sat_reg_going_drivers.append(shuffled_drivers[i])
                sat_early_leaving_drivers.append(shuffled_drivers[i])
                print("Leftover Driver:", shuffled_drivers[i], "|", cars_going[shuffled_drivers[i]])
    if(len(leftover) > 0):
        for k in range(len(leftover)):
            maven_sat.append(leftover[k])
    if(len(maven_sat) > 0):
        sat_reg_going_drivers.append(leftover[0])
        sat_early_leaving_drivers.append(leftover[0])
        print("Saturday Maven Car:", maven_sat)
    print("\n")

    #Print The Cars Going There
    print(sat_reg_going_drivers)
    print(sat_KG_going_drivers)
    print(sat_reg_leaving_drivers)
    print(sat_early_leaving_drivers)
    print("\n")
    sat_cars_going_reg = {}
    sat_cars_going_KG = {}
    sat_cars_leaving_reg = {}
    sat_cars_leaving_early = {}
    for i in cars_going:
        for j in sat_reg_going_drivers:
            if(i == j):
                sat_cars_going_reg[j] = cars_going[i]
        for j in sat_KG_going_drivers:
            if(i == j):
                sat_cars_going_KG[j] = cars_going[i]
        for j in sat_reg_leaving_drivers:
            if(i == j):
                sat_cars_leaving_reg[j] = cars_going[i]
        for j in sat_early_leaving_drivers:
            if(i == j):
                sat_cars_leaving_early[j] = cars_going[i]

    #SUNDAY CARS
    sun_cars_going = {}
    sun_leftover = []  # Leftover people is a combination of people who don't have a spot,
    # and will likely have to discuss among the car what they want to do if they want to leave early or not
    maven_sun = []

    sun_reg_going_drivers = []
    sun_KG_going_drivers = []
    sun_reg_leaving_drivers = []
    sun_early_leaving_drivers = []

    # Saturday Regular
    for i in range(len(shuffled_drivers)):
        if (shuffled_drivers[i] in attendanceList[4]):
            sun_cars_going[shuffled_drivers[i]] = []
            attendanceList[4].remove(shuffled_drivers[i])
            if (len(attendanceList[4]) < 4):
                for j in range(len(attendanceList[4])):
                    choice = random.choice(attendanceList[4])
                    if (choice not in shuffled_drivers):
                        sun_cars_going[shuffled_drivers[i]].append(choice)
                        attendanceList[4].remove(choice)
            else:
                for j in range(4):
                    while (True):
                        choice = random.choice(attendanceList[4])
                        if (choice not in shuffled_drivers):
                            break
                    sun_cars_going[shuffled_drivers[i]].append(choice)
                    attendanceList[4].remove(choice)
            if (len(sun_cars_going[shuffled_drivers[i]]) > 1):
                sun_reg_going_drivers.append(shuffled_drivers[i])
                sun_reg_leaving_drivers.append(shuffled_drivers[i])
                print("Sunday Driver:", shuffled_drivers[i], "|", sun_cars_going[shuffled_drivers[i]])
    if (len(attendanceList[4]) > 0):
        for k in range(len(attendanceList[4])):
            sun_leftover.append(attendanceList[4][k])

    # Saturday Early
    for i in range(len(shuffled_drivers)):
        if (shuffled_drivers[i] in attendanceList[5]):
            sun_cars_going[shuffled_drivers[i]] = []
            attendanceList[5].remove(shuffled_drivers[i])
            if (len(attendanceList[5]) < 4):
                for j in range(len(attendanceList[5])):
                    choice = random.choice(attendanceList[5])
                    if (choice not in shuffled_drivers):
                        sun_cars_going[shuffled_drivers[i]].append(choice)
                        attendanceList[5].remove(choice)
            else:
                for j in range(4):
                    while (True):
                        choice = random.choice(attendanceList[5])
                        if (choice not in shuffled_drivers):
                            break
                    sun_cars_going[shuffled_drivers[i]].append(choice)
                    attendanceList[5].remove(choice)
            if (len(sun_cars_going[shuffled_drivers[i]]) > 1):
                sun_reg_going_drivers.append(shuffled_drivers[i])
                sun_early_leaving_drivers.append(shuffled_drivers[i])
                print("Sunday Early Driver:", shuffled_drivers[i], "|", sun_cars_going[shuffled_drivers[i]])
    if (len(attendanceList[5]) > 0):
        for k in range(len(attendanceList[5])):
            sun_leftover.append(attendanceList[5][k])

    # Saturday KG
    for i in range(len(shuffled_drivers)):
        if (shuffled_drivers[i] in attendanceList[6]):
            sun_cars_going[shuffled_drivers[i]] = []
            attendanceList[6].remove(shuffled_drivers[i])
            if (len(attendanceList[6]) < 4):
                for j in range(len(attendanceList[6])):
                    choice = random.choice(attendanceList[6])
                    if (choice not in shuffled_drivers):
                        sun_cars_going[shuffled_drivers[i]].append(choice)
                        attendanceList[6].remove(choice)
            else:
                for j in range(4):
                    while (True):
                        choice = random.choice(attendanceList[6])
                        if (choice not in shuffled_drivers):
                            break
                    sun_cars_going[shuffled_drivers[i]].append(choice)
                    attendanceList[6].remove(choice)
            if (len(sun_cars_going[shuffled_drivers[i]]) > 1):
                sun_KG_going_drivers.append(shuffled_drivers[i])
                sun_reg_leaving_drivers.append(shuffled_drivers[i])
                print("Sunday KG Driver:", shuffled_drivers[i], "|", sun_cars_going[shuffled_drivers[i]])
    if (len(attendanceList[6]) > 0):
        for k in range(len(attendanceList[6])):
            sun_leftover.append(attendanceList[6][k])

    # Saturday KG/Early
    for i in range(len(shuffled_drivers)):
        if (shuffled_drivers[i] in attendanceList[7]):
            sun_cars_going[shuffled_drivers[i]] = []
            attendanceList[7].remove(shuffled_drivers[i])
            if (len(attendanceList[7]) < 4):
                for j in range(len(attendanceList[7])):
                    choice = random.choice(attendanceList[7])
                    if (choice not in shuffled_drivers):
                        sun_cars_going[shuffled_drivers[i]].append(choice)
                        attendanceList[7].remove(choice)
            else:
                for j in range(4):
                    while (True):
                        choice = random.choice(attendanceList[7])
                        if (choice not in shuffled_drivers):
                            break
                    sun_cars_going[shuffled_drivers[i]].append(choice)
                    attendanceList[7].remove(choice)
            if (len(sun_cars_going[shuffled_drivers[i]]) > 1):
                sun_KG_going_drivers.append(shuffled_drivers[i])
                sun_early_leaving_drivers.append(shuffled_drivers[i])
                print("Sunday KG Early Driver:", shuffled_drivers[i], "|", sun_cars_going[shuffled_drivers[i]])
    if (len(attendanceList[7]) > 0):
        for k in range(len(attendanceList[7])):
            sun_leftover.append(attendanceList[7][k])

    for i in sun_cars_going:
        if (len(sun_cars_going[i]) == 0):
            sun_leftover.append(i)
        if (len(sun_cars_going[i]) == 1):
            sun_leftover.append(i)
            for k in sun_cars_going[i]:
                sun_leftover.append(k)

    for i in range(len(shuffled_drivers)):
        if (shuffled_drivers[i] in sun_leftover):
            sun_cars_going[shuffled_drivers[i]] = []
            sun_leftover.remove(shuffled_drivers[i])
            if (len(sun_leftover) < 4):
                for j in range(len(sun_leftover)):
                    choice = random.choice(sun_leftover)
                    if (choice not in shuffled_drivers):
                        sun_cars_going[shuffled_drivers[i]].append(choice)
                        sun_leftover.remove(choice)
            else:
                for j in range(4):
                    while (True):
                        choice = random.choice(sun_leftover)
                        if (choice not in shuffled_drivers):
                            break
                    sun_cars_going[shuffled_drivers[i]].append(choice)
                    sun_leftover.remove(choice)
            if (len(sun_cars_going[shuffled_drivers[i]]) > 1):
                sun_reg_going_drivers.append(shuffled_drivers[i])
                sun_early_leaving_drivers.append(shuffled_drivers[i])
                print("Leftover Driver:", shuffled_drivers[i], "|", sun_cars_going[shuffled_drivers[i]])
    if (len(sun_leftover) > 0):
        for k in range(len(sun_leftover)):
            maven_sun.append(sun_leftover[k])
    if (len(maven_sun) > 0):
        sun_reg_going_drivers.append(sun_leftover[0])
        sun_early_leaving_drivers.append(sun_leftover[0])
        print("Sunday Maven Car:", maven_sun)
    print("\n")

    # Print The Cars Going There
    print(sun_reg_going_drivers)
    print(sun_KG_going_drivers)
    print(sun_reg_leaving_drivers)
    print(sun_early_leaving_drivers)
    print("\n")
    sun_cars_going_reg = {}
    sun_cars_going_KG = {}
    sun_cars_leaving_reg = {}
    sun_cars_leaving_early = {}
    for i in sun_cars_going:
        for j in sun_reg_going_drivers:
            if (i == j):
                sun_cars_going_reg[j] = sun_cars_going[i]
        for j in sun_KG_going_drivers:
            if (i == j):
                sun_cars_going_KG[j] = sun_cars_going[i]
        for j in sun_reg_leaving_drivers:
            if (i == j):
                sun_cars_leaving_reg[j] = sun_cars_going[i]
        for j in sun_early_leaving_drivers:
            if (i == j):
                sun_cars_leaving_early[j] = sun_cars_going[i]

    print("Saturday Regular Going:", sat_cars_going_reg)
    print("Saturday KG Going:", sat_cars_going_KG)
    print("Saturday Regular Leaving:", sat_cars_leaving_reg)
    print("Saturday Early Leaving:", sat_cars_leaving_early)
    print("\n")
    print("Sunday Regular Going:", sun_cars_going_reg)
    print("Sunday KG Going:", sun_cars_going_KG)
    print("Sunday Regular Leaving:", sun_cars_leaving_reg)
    print("Sunday Early Leaving:", sun_cars_leaving_early)
    print("\n")

    return("Saturday Regular Going:", sat_cars_going_reg,
           "\nSaturday KG Going:", sat_cars_going_KG,
           "\nSaturday Regular Leaving:", sat_cars_leaving_reg,
           "\nSaturday Early Leaving:", sat_cars_leaving_early,
           "\n",
           "\nSunday Regular Going:", sun_cars_going_reg,
           "\nSunday KG Going:", sun_cars_going_KG,
           "\nSunday Regular Leaving:", sun_cars_leaving_reg,
           "\nSunday Early Leaving:", sun_cars_leaving_early)

global today
global day
global nextsat
global nextsun
def what_fucking_day_is_it():
    today = date.today()
    day = datetime.datetime.today().weekday()
    week_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    print(week_days[day])
    print("Today's date:", today)
    print(day)
    if(day == 5 or day == 6):
        print("Today is Weekend")
    else:
        print("Today is Weekday")
        nextsat = today + datetime.timedelta(5-day)
        nextsun = today + datetime.timedelta(6-day)
        print("Next Saturday is: ", nextsat)
        print("Next Sunday is: ", nextsun)
def judgement_time():
    current_time = time.strftime("%H:%M:%S")
    post_time = "00:00:00"
    nextfri = today + datetime.timedelta(4-day)
    if(today == nextfri and current_time == post_time):
        Brianna_Wants_To_Be_Emailed_Instead()
def Brianna_Wants_To_Be_Emailed_Instead():
    what_fucking_day_is_it()
    make_cars(nextsat, nextsun)
    # creates SMTP session
    s = smtplib.SMTP('smtp.gmail.com', 587)

    # start TLS for security
    s.ehlo()
    s.starttls()
    s.ehlo()

    # Authentication
    s.login("usc.dragonboat@gmail.com", "fightforgold")

    # message to be sent
    message = "Hello Brianna, \n    This is a test"
    counter = 0
    message = []
    message.append("List of Drivers: ")
    while(counter < len(drivers)):
        ##people = ""
        ##for driver in drivers:
        ##    people += driver + " "
        ##message.append("List of Drivers: " + people)
        message.append(drivers[counter])
        counter+=1
    print(message)
    # sending the mail
    #s.sendmail("edisonsiu8899@gmail.com", "sunb@usc.edu", message)
    #s.sendmail("usc.dragonboat@gmail.com", "edisonsiu8899@gmail.com", message)
    s.sendmail("usc.dragonboat@gmail.com", "sunb@usc.edu", message)

    # terminating the session
    s.quit()

def main():
    #make_cars('2020-02-08', '2020-02-09')
    #what_fucking_day_is_it()
    #Brianna_Wants_To_Be_Emailed_Instead()
    judgement_time()
main()

